# -*- coding: utf-8 -*-
"""
Created on Wed Feb 24 22:02:20 2016

@author: pkaran
"""

from nltk.stem.lancaster import LancasterStemmer
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.stem import WordNetLemmatizer
from nltk.corpus import stopwords
lancaster_stemmer = LancasterStemmer()
wordnet_lemmatizer = WordNetLemmatizer()

import openpyxl
wb2 = openpyxl.load_workbook('training_TM.xlsx')
ws2 = wb2.get_sheet_by_name('Sheet1')

for i in range(2, 22957):
     count = 0
     a = 0.0
     b = 0.0
     c = 0.0
	
     f_essay = ws2.cell(row = i, column = 1)
     review = f_essay.value
	
     yelp_tokanization_stopwords = [w for w in review if not w in stopwords.words("english")]
     yelp_tokanization_stopwords1= [w for w in review if not w in stopwords.words("english")]
     #print(yelp_tokanization_stopwords)
     yelp_tokanization_sentence = sent_tokenize(review)
     #print(yelp_tokanization_sentence)
     yelp_tokanization_word = word_tokenize(review)
     #print(yelp_tokanization_word)
     stem = [lancaster_stemmer.stem(w) for w in yelp_tokanization_stopwords]
     yelp_tokanization_lemma = [wordnet_lemmatizer.lemmatize(w) for w in yelp_tokanization_stopwords]
    	
     count = set(yelp_tokanization_stopwords1).intersection(yelp_tokanization_lemma)
     a = len(count) * 100
     b = len(yelp_tokanization_stopwords1)
     c = a / b
     #print(c)
	
     ws2.cell(row = i, column = 36).value = c
wb2.save('training_TM.xlsx')

